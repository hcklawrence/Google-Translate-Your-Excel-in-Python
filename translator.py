from mtranslate import translate

def toChinese(orignal):
    message = translate(orignal, 'zh-cn')
    return message

from openpyxl import Workbook, load_workbook

inputExcelFile  = 'translate.xlsx'
outputExcelFile = 'output.xlsx'

theColumnYouWant2Translate              = 2
theColumnYouWant2PlaceThoseTranslation  = 3

workbook = load_workbook(inputExcelFile)
sheets = workbook.sheetnames
booksheet = workbook[sheets[0]]

rows = booksheet.rows
cols = booksheet.columns

newbook = []
for row in rows:
    line = [col.value for col in row]
    newbook.append(line)

for i in range(1,len(newbook)):
    orignal = newbook[i][theColumnYouWant2Translate]
    message = toChinese(orignal)
    newbook[i][theColumnYouWant2PlaceThoseTranslation] = message
    print(str(i) + " " + message)

print('translation done')

import pandas as pd
pd.DataFrame(newbook).to_excel(outputExcelFile, header = False, index = False)